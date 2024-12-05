import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from .models import User


def export_users_xlsx(request):
    # Create a workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    # Add header row
    headers = ['Username', 'First Name', 'Last Name', 'Email Address', 'CV']
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font

    # Add user data
    rows = User.objects.all().values_list('username', 'first_name', 'last_name', 'email', 'cv')
    for row_num, row in enumerate(rows, start=2):
        for col_num, cell_value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_num).value = cell_value

    # Prepare the response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="users.xlsx"'

    # Save the workbook to the response
    wb.save(response)
    return response
